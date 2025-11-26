from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from django.forms import inlineformset_factory
from decimal import Decimal
import datetime
import logging
from openpyxl import load_workbook

from .models import Calificacion, CalificacionFactor
from .forms import CalificacionForm, CalificacionFactorForm, ExcelUploadForm
from .kafka_producer import enviar_evento_calificacion   # 👉 integración Kafka

# 👉 métricas Prometheus
from miapp.views_monitor import (
    login_counter,
    calificacion_creada_counter,
    calificacion_actualizada_counter,
    calificacion_importada_counter,
)

logger = logging.getLogger('miapp')


def login_view(request):
    if request.user.is_authenticated:
        return redirect('login')
    if request.method == 'POST':
        user = authenticate(
            request,
            username=request.POST.get('username'),
            password=request.POST.get('password')
        )
        if user is not None:
            login(request, user)
            login_counter.inc()
            return redirect('lista_calificaciones')
        messages.error(request, 'Usuario o contraseña incorrectos')
    return render(request, 'login.html')


def logout_view(request):
    logout(request)
    return redirect('login')


@login_required(login_url='login')
def lista_calificaciones(request):
    calificaciones = Calificacion.objects.prefetch_related('factores').all()
    return render(request, 'calificaciones_login.html', {'calificaciones': calificaciones})


@login_required(login_url='login')
def crear_calificacion(request):
    FactorFormSet = inlineformset_factory(
        Calificacion,
        CalificacionFactor,
        form=CalificacionFactorForm,
        extra=31,
        can_delete=False
    )

    if request.method == 'POST':
        form = CalificacionForm(request.POST)
        formset = FactorFormSet(request.POST)
        if form.is_valid() and formset.is_valid():
            calificacion = form.save()
            factores = formset.save(commit=False)
            for f in factores:
                if f.valor_factor is not None:
                    f.calificacion = calificacion
                    f.save()

            enviar_evento_calificacion(calificacion)
            calificacion_creada_counter.inc()

            messages.success(request, 'Calificación creada exitosamente')
            return redirect('lista_calificaciones')
    else:
        form = CalificacionForm()
        initial_data = [{'numero_factor': i} for i in range(8, 39)]
        formset = FactorFormSet(queryset=CalificacionFactor.objects.none(), initial=initial_data)

    return render(request, 'calificacion_form.html', {'form': form, 'formset': formset})


@login_required(login_url='login')
def actualizar_calificacion(request, id_calificacion):
    calificacion = get_object_or_404(Calificacion, id_calificacion=id_calificacion)
    FactorFormSet = inlineformset_factory(
        Calificacion,
        CalificacionFactor,
        form=CalificacionFactorForm,
        extra=0,
        can_delete=True
    )
    if request.method == 'POST':
        form = CalificacionForm(request.POST, instance=calificacion)
        formset = FactorFormSet(request.POST, instance=calificacion)
        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()

            enviar_evento_calificacion(calificacion)
            calificacion_actualizada_counter.inc()

            messages.success(request, 'Calificación actualizada exitosamente')
            return redirect('lista_calificaciones')
    else:
        form = CalificacionForm(instance=calificacion)
        existentes = calificacion.factores.values_list('numero_factor', flat=True)
        initial_data = [{'numero_factor': i} for i in range(8, 39) if i not in existentes]
        formset = FactorFormSet(instance=calificacion, initial=initial_data)

    return render(request, 'calificacion_form.html', {'form': form, 'formset': formset})


@login_required(login_url='login')
def eliminar_calificacion(request, id_calificacion):
    calificacion = get_object_or_404(Calificacion, id_calificacion=id_calificacion)
    calificacion.delete()
    messages.success(request, 'Calificación eliminada exitosamente')
    return redirect('lista_calificaciones')


@login_required(login_url='login')
def api_calificaciones(request):
    data = list(Calificacion.objects.values())
    return JsonResponse(data, safe=False)


@login_required(login_url='login')
def importar_calificaciones(request):
    def to_decimal(x, places=None):
        if x is None or (isinstance(x, str) and not x.strip()):
            return None
        try:
            if isinstance(x, str):
                x = x.strip().replace('.', '').replace(',', '.')
            d = Decimal(str(x))
            if places is not None:
                q = Decimal('1.' + '0'*places)
                d = d.quantize(q)
            return d
        except Exception:
            return None

    def to_date(x):
        if isinstance(x, (datetime.date, datetime.datetime)):
            return x.date() if isinstance(x, datetime.datetime) else x
        xs = str(x).strip().replace('.', '-').replace('/', '-')
        for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%m-%d-%Y'):
            try:
                return datetime.datetime.strptime(xs, fmt).date()
            except Exception:
                continue
        return None

    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            file_obj = form.cleaned_data['archivo']
            try:
                wb = load_workbook(file_obj, data_only=True)

                # 👉 Hoja 1: Calificaciones
                ws_cal = wb['Calificaciones']   # usa el nombre real de la hoja
                headers_cal = [str(c.value).strip().lower().replace(" ", "_")
                               for c in next(ws_cal.iter_rows(min_row=1, max_row=1))]

                # 👉 Hoja 2: Factores
                ws_fac = wb['Factores']         # usa el nombre real de la hoja
                headers_fac = [str(c.value).strip().lower().replace(" ", "_")
                               for c in next(ws_fac.iter_rows(min_row=1, max_row=1))]

            except Exception as e:
                logger.error(f"No se pudo leer el Excel: {e}")
                messages.error(request, "No se pudo leer el Excel.")
                return redirect('lista_calificaciones')

            # 👉 Importar calificaciones
            cal_map = {}
            for row in ws_cal.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers_cal, row))
                calificacion = Calificacion.objects.create(
                    instrumento=(row_dict.get('instrumento') or '').strip(),
                    descripcion=row_dict.get('descripcion'),
                    fecha_pago=to_date(row_dict.get('fecha_pago')),
                    valor_historico=to_decimal(row_dict.get('valor_historico'))
                )
                cal_map[row_dict.get('id_calificacion')] = calificacion

            # 👉 Importar factores y relacionarlos
            for row in ws_fac.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers_fac, row))
                cal_id = row_dict.get('calificacion_id')
                calificacion = cal_map.get(cal_id)
                if calificacion:
                    CalificacionFactor.objects.create(
                        calificacion=calificacion,
                        numero_factor=row_dict.get('numero_factor'),
                        valor_factor=to_decimal(row_dict.get('valor_factor'), places=8) or Decimal('0')
                    )

            messages.success(request, "Se importaron calificaciones y factores desde Excel.")
            return redirect('lista_calificaciones')
    else:
        form = ExcelUploadForm()

    return render(request, 'importar_calificaciones.html', {'form': form})
